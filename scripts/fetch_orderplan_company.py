"""
나라장터 발주계획현황 (기술용역) 데이터 수집 및 정렬 스크립트 v2
- 엔드포인트: /ao/OrderPlanSttusService/getOrderPlanSttusListServcPPSSrch
- 키워드 그룹 순서 → 금액 내림차순 정렬
"""

import os
import sys
import requests
import pandas as pd
from datetime import datetime, timedelta
import logging

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# ── 환경변수 ──────────────────────────────────────────────────
API_KEY          = os.environ["NARA_API_KEY"]
TELEGRAM_TOKEN   = os.environ["TELEGRAM_BOT_TOKEN"]
TELEGRAM_CHAT_ID = os.environ["TELEGRAM_CHAT_ID"]

# ── 상수 ──────────────────────────────────────────────────────
BASE_URL = (
    "http://apis.data.go.kr/1230000/ao/OrderPlanSttusService"
    "/getOrderPlanSttusListServcPPSSrch"
)
PAGE_SIZE = 999

# ★ 키워드 순서 중요 (앞 키워드 우선 배치, 중복 시 첫 번째 그룹에만 포함)
KEYWORDS = ["타당성", "기본구상", "기본계획", "설계", "건설사업관리"]

# ★ 키워드 별칭: 본 키워드 미매칭 시 해당 그룹으로 분류할 추가 단어
KEYWORD_ALIASES = {
    "기본계획": [
        "관리계획", "재정비", "지구단위계획", "개발계획",
        "지구지정", "구역지정", "환지처분", "도시계획",
    ],
}

COLUMN_MAP = {
    "bizNm":          "사업명",
    "orderInsttNm":   "발주기관",
    "totlmngInsttNm": "총괄기관",
    "jrsdctnDivNm":   "소관구분",
    "sumOrderAmt":    "합계발주금액(원)",
    "orderYear":      "발주년도",
    "orderMnth":      "발주월",
    "cnsttyDivNm":    "공종구분",
    "cntrctMthdNm":   "계약방법",
    "prcrmntMethd":   "조달방식",
    "bsnsTyNm":       "업무유형",
    "nticeDt":        "게시일시",
    "deptNm":         "담당부서",
    "ofclNm":         "담당자",
    "telNo":          "전화번호",
    "bidNtceNoList":  "공고번호",
}


def get_target_date_range() -> tuple[str, str, str]:
    target = os.environ.get("TARGET_DATE", "").strip()
    if target and len(target) == 8:
        base = datetime.strptime(target, "%Y%m%d")
    else:
        from datetime import timezone
        KST = timezone(timedelta(hours=9))
        base = datetime.now(KST).replace(tzinfo=None) - timedelta(days=1)

    start    = base.strftime("%Y%m%d") + "0000"
    end      = base.strftime("%Y%m%d") + "2359"
    date_str = base.strftime("%Y%m%d")
    return start, end, date_str


def fetch_all_pages(start_dt: str, end_dt: str) -> list[dict]:
    all_items = []
    page = 1

    while True:
        url = (
            f"{BASE_URL}"
            f"?ServiceKey={API_KEY}"
            f"&pageNo={page}"
            f"&numOfRows={PAGE_SIZE}"
            f"&type=json"
            f"&inqryBgnDt={start_dt}"
            f"&inqryEndDt={end_dt}"
        )

        try:
            resp = requests.get(url, timeout=30)
            logger.info(f"  HTTP {resp.status_code} (page {page})")

            if resp.status_code != 200:
                logger.error(f"API 오류: {resp.text[:300]}")
                break

            data = resp.json()

        except Exception as e:
            logger.error(f"API 호출 실패 (page {page}): {e}")
            break

        try:
            body      = data["response"]["body"]
            total_cnt = int(body.get("totalCount", 0))
            items     = body.get("items", {})

            if isinstance(items, dict):
                row_list = items.get("item", [])
                if isinstance(row_list, dict):
                    row_list = [row_list]
            elif isinstance(items, list):
                row_list = items
            else:
                row_list = []

            all_items.extend(row_list)
            logger.info(
                f"  page {page}: {len(row_list)}건 "
                f"(누계 {len(all_items)}/{total_cnt})"
            )

            if len(all_items) >= total_cnt or len(row_list) == 0:
                break
            page += 1

        except (KeyError, TypeError) as e:
            logger.error(f"응답 파싱 오류: {e}")
            logger.error(f"원본: {str(data)[:500]}")
            break

    return all_items


def assign_keyword_group(name: str) -> str:
    """사업명에서 첫 번째 매칭 키워드 반환
    1단계: KEYWORDS 본 키워드 순서대로 매칭
    2단계: KEYWORD_ALIASES 별칭으로 매칭 (본 키워드 미매칭 시에만)
    """
    for kw in KEYWORDS:
        if kw in name:
            return kw
    for group, aliases in KEYWORD_ALIASES.items():
        for alias in aliases:
            if alias in name:
                return group
    return ""


def build_dataframe(items: list[dict]) -> pd.DataFrame:
    """DataFrame 변환 → 키워드 필터·태깅 → 그룹순서+금액 내림차순 정렬"""
    if not items:
        return pd.DataFrame()

    df = pd.DataFrame(items)

    for col in COLUMN_MAP:
        if col not in df.columns:
            df[col] = ""

    df = df[list(COLUMN_MAP.keys())].rename(columns=COLUMN_MAP)

    # 금액 숫자 변환
    df["합계발주금액(원)"] = (
        pd.to_numeric(df["합계발주금액(원)"], errors="coerce")
        .fillna(0).astype(int)
    )

    # ★ 키워드 태깅
    df["검색키워드"] = df["사업명"].apply(assign_keyword_group)

    # ★ 키워드 없는 행 제거
    df = df[df["검색키워드"] != ""].copy()

    # ★ 중복 공고 제거: 동일 공고번호 중 첫 번째 키워드 그룹만 유지
    kw_order = {kw: i for i, kw in enumerate(KEYWORDS)}
    df["_kw_order"] = df["검색키워드"].map(kw_order)
    df = df.sort_values(["공고번호", "_kw_order"])
    df = df.drop_duplicates(subset="공고번호", keep="first")
    df = df.drop(columns=["_kw_order"])

    # ★ 그룹 순서 → 금액 내림차순 정렬
    df["_group_order"] = df["검색키워드"].map(kw_order)
    df = df.sort_values(
        ["_group_order", "합계발주금액(원)"],
        ascending=[True, False]
    ).reset_index(drop=True)
    df = df.drop(columns=["_group_order"])
    df.index += 1

    # ★ 검색키워드 컬럼을 앞으로 이동
    cols = ["검색키워드"] + [c for c in df.columns if c != "검색키워드"]
    df = df[cols]

    # 금액 천단위 콤마
    df["합계발주금액(원)"] = df["합계발주금액(원)"].apply(lambda x: f"{x:,}")

    # 키워드별 건수 로그
    for kw in KEYWORDS:
        cnt = (df["검색키워드"] == kw).sum()
        logger.info(f"  {kw}: {cnt}건")
    logger.info(f"최종 합계: {len(df)}건")

    return df


def save_excel(df: pd.DataFrame, date_str: str) -> str:
    filename = f"나라장터(회사용)_기술용역_발주계획_{date_str}.xlsx"
    filepath = f"/tmp/{filename}"

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(
            writer, sheet_name="기술용역_발주계획",
            index=True, index_label="순위"
        )
        ws = writer.sheets["기술용역_발주계획"]

        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value else 0) for c in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(
                max_len + 4, 60
            )

        from openpyxl.styles import PatternFill, Font, Alignment
        fill = PatternFill("solid", fgColor="1B5E20")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")

    logger.info(f"Excel 저장: {filepath}")
    return filepath


def send_telegram_message(text: str):
    requests.post(
        f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
        json={
            "chat_id":    TELEGRAM_CHAT_ID,
            "text":       text,
            "parse_mode": "Markdown"
        },
        timeout=15,
    )


def send_telegram_file(filepath: str, date_str: str, df: pd.DataFrame):
    y, m, d = date_str[:4], date_str[4:6], date_str[6:]

    # ★ 키워드별 건수 요약
    kw_summary = "\n".join(
        f"  • {kw}: {(df['검색키워드'] == kw).sum()}건"
        for kw in KEYWORDS
    )

    msg = (
        f"📌 *나라장터 기술용역 발주계획 (회사용)*\n"
        f"📅 기준일: {y}-{m}-{d}\n"
        f"📊 총 수집건수: *{len(df)}건*\n"
        f"\n{kw_summary}\n"
        f"\n🔽 키워드 그룹 순서 → 금액 내림차순 정렬"
    )
    send_telegram_message(msg)

    with open(filepath, "rb") as f:
        resp = requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
            data={"chat_id": TELEGRAM_CHAT_ID},
            files={"document": (os.path.basename(filepath), f)},
            timeout=60,
        )

    if resp.status_code == 200:
        logger.info("텔레그램 전송 성공 ✅")
    else:
        logger.error(f"텔레그램 전송 실패: {resp.text}")


def main():
    start_dt, end_dt, date_str = get_target_date_range()
    logger.info(f"▶ 발주계획 수집 시작: {date_str} (00:00 ~ 23:59)")

    items = fetch_all_pages(start_dt, end_dt)
    logger.info(f"전체 수집: {len(items)}건")

    if not items:
        y, m, d = date_str[:4], date_str[4:6], date_str[6:]
        send_telegram_message(
            f"📌 *나라장터 기술용역 발주계획 (회사용)*\n"
            f"📅 기준일: {y}-{m}-{d}\n"
            f"ℹ️ 해당일 등록 데이터가 없습니다."
        )
        return

    df = build_dataframe(items)

    if df.empty:
        y, m, d = date_str[:4], date_str[4:6], date_str[6:]
        send_telegram_message(
            f"📌 *나라장터 기술용역 발주계획 (회사용)*\n"
            f"📅 기준일: {y}-{m}-{d}\n"
            f"ℹ️ 키워드 해당 데이터가 없습니다.\n"
            f"🔍 검색어: {', '.join(KEYWORDS)}"
        )
        return

    filepath = save_excel(df, date_str)
    send_telegram_file(filepath, date_str, df)
    logger.info("▶ 완료")


if __name__ == "__main__":
    main()
