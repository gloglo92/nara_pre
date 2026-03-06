"""
나라장터 사전규격공개 (기술용역) 데이터 수집 및 정렬 스크립트 v4
- 엔드포인트: /ao/HrcspSsstndrdInfoService/getPublicPrcureThngInfoServc
- 공식 참고문서(조달청_OpenAPI참고자료) 기준으로 작성
- 매일 전일 데이터를 수집하여 Excel로 저장 후 텔레그램 발송
"""

import os
import sys
import requests
import pandas as pd
from datetime import datetime, timedelta
import logging

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# ── 환경변수 ──────────────────────────────────────────────────
API_KEY          = os.environ["NARA_API_KEY"]
TELEGRAM_TOKEN   = os.environ["TELEGRAM_BOT_TOKEN"]
TELEGRAM_CHAT_ID = os.environ["TELEGRAM_CHAT_ID"]

# ── 상수 ──────────────────────────────────────────────────────
# ★ 공식 문서 확인된 정확한 엔드포인트
BASE_URL  = (
    "http://apis.data.go.kr/1230000/ao/HrcspSsstndrdInfoService"
    "/getPublicPrcureThngInfoServc"
)
PAGE_SIZE = 999

# ★ 사업명 키워드 필터 (이 단어 중 하나라도 포함된 건만 추출)
KEYWORDS = ["타당성", "기본계획", "설계", "건설사업관리"]

# 응답 필드 → 한글 컬럼 매핑 (공식 문서 응답 필드 기준)
COLUMN_MAP = {
    "bfSpecRgstNo":    "사전규격등록번호",
    "prdctClsfcNoNm":  "사업명(품명)",
    "orderInsttNm":    "발주기관",
    "rlDminsttNm":     "수요기관",
    "asignBdgtAmt":    "배정예산액(원)",
    "rcptDt":          "등록일시",
    "opninRgstClseDt": "의견등록마감일",
    "dlvrTmlmtDt":     "납품기한",
    "swBizObjYn":      "SW사업여부",
    "bidNtceNoList":   "공고번호",
    "refNo":           "참조번호",
    "specDocFileUrl1": "규격서URL",
}


def get_target_date_range() -> tuple[str, str, str]:
    """전일 00:00 ~ 23:59 범위 반환"""
    target = os.environ.get("TARGET_DATE", "").strip()
    if target and len(target) == 8:
        base = datetime.strptime(target, "%Y%m%d")
    else:
        base = datetime.now() - timedelta(days=1)

    start    = base.strftime("%Y%m%d") + "0000"
    end      = base.strftime("%Y%m%d") + "2359"
    date_str = base.strftime("%Y%m%d")
    return start, end, date_str


def fetch_all_pages(start_dt: str, end_dt: str) -> list[dict]:
    """페이징 처리하여 전체 데이터 수집"""
    all_items = []
    page = 1

    while True:
        # URL 직접 조합 (인증키 이중인코딩 방지)
        url = (
            f"{BASE_URL}"
            f"?ServiceKey={API_KEY}"   # ★ 공식문서 기준 대문자 S
            f"&pageNo={page}"
            f"&numOfRows={PAGE_SIZE}"
            f"&type=json"
            f"&inqryDiv=1"             # 1=등록일시 기준 조회
            f"&inqryBgnDt={start_dt}"
            f"&inqryEndDt={end_dt}"
        )

        try:
            resp = requests.get(url, timeout=30)
            logger.info(f"  HTTP {resp.status_code} (page {page})")

            if resp.status_code != 200:
                logger.error(f"API 오류: {resp.text[:300]}")
                break

            data = resp.json()

        except Exception as e:
            logger.error(f"API 호출 실패 (page {page}): {e}")
            break

        # 응답 파싱
        try:
            body      = data["response"]["body"]
            total_cnt = int(body.get("totalCount", 0))
            items     = body.get("items", {})

            if isinstance(items, dict):
                row_list = items.get("item", [])
                if isinstance(row_list, dict):
                    row_list = [row_list]
            elif isinstance(items, list):
                row_list = items
            else:
                row_list = []

            all_items.extend(row_list)
            logger.info(
                f"  page {page}: {len(row_list)}건 "
                f"(누계 {len(all_items)}/{total_cnt})"
            )

            if len(all_items) >= total_cnt or len(row_list) == 0:
                break
            page += 1

        except (KeyError, TypeError) as e:
            logger.error(f"응답 파싱 오류: {e}")
            logger.error(f"원본: {str(data)[:500]}")
            break

    return all_items


def filter_by_keywords(df: pd.DataFrame) -> pd.DataFrame:
    """사업명에 키워드가 포함된 행만 필터링"""
    if df.empty:
        return df
    pattern = "|".join(KEYWORDS)
    mask = df["사업명(품명)"].str.contains(pattern, na=False)
    filtered = df[mask].reset_index(drop=True)
    filtered.index += 1
    logger.info(
        f"키워드 필터링 ({', '.join(KEYWORDS)}): "
        f"{len(df)}건 → {len(filtered)}건"
    )
    return filtered


def build_dataframe(items: list[dict]) -> pd.DataFrame:
    """DataFrame 변환 및 예산액 내림차순 정렬"""
    if not items:
        return pd.DataFrame()

    df = pd.DataFrame(items)

    for col in COLUMN_MAP:
        if col not in df.columns:
            df[col] = ""

    df = df[list(COLUMN_MAP.keys())].rename(columns=COLUMN_MAP)

    # 예산액 숫자 변환 → 정렬 → 천단위 콤마
    df["배정예산액(원)"] = (
        pd.to_numeric(df["배정예산액(원)"], errors="coerce")
        .fillna(0).astype(int)
    )
    df = df.sort_values("배정예산액(원)", ascending=False).reset_index(drop=True)
    df.index += 1
    df["배정예산액(원)"] = df["배정예산액(원)"].apply(lambda x: f"{x:,}")

    return df


def save_excel(df: pd.DataFrame, date_str: str) -> str:
    """Excel 저장 (헤더 스타일 포함)"""
    filename = f"나라장터_기술용역_사전규격_{date_str}.xlsx"
    filepath = f"/tmp/{filename}"

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(
            writer, sheet_name="기술용역_사전규격",
            index=True, index_label="순위"
        )
        ws = writer.sheets["기술용역_사전규격"]

        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) if c.value else 0) for c in col_cells
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(
                max_len + 4, 60
            )

        from openpyxl.styles import PatternFill, Font, Alignment
        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")

    logger.info(f"Excel 저장: {filepath}")
    return filepath


def send_telegram_message(text: str):
    requests.post(
        f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage",
        json={
            "chat_id":    TELEGRAM_CHAT_ID,
            "text":       text,
            "parse_mode": "Markdown"
        },
        timeout=15,
    )


def send_telegram_file(filepath: str, date_str: str, count: int):
    y, m, d = date_str[:4], date_str[4:6], date_str[6:]
    msg = (
        f"📋 *나라장터 기술용역 사전규격공개*\n"
        f"📅 기준일: {y}-{m}-{d}\n"
        f"📊 수집건수: *{count}건*\n"
        f"🔍 필터: {', '.join(KEYWORDS)}\n"
        f"🔽 배정예산액 높은 순 정렬"
    )
    send_telegram_message(msg)

    with open(filepath, "rb") as f:
        resp = requests.post(
            f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument",
            data={"chat_id": TELEGRAM_CHAT_ID},
            files={"document": (os.path.basename(filepath), f)},
            timeout=60,
        )

    if resp.status_code == 200:
        logger.info("텔레그램 전송 성공 ✅")
    else:
        logger.error(f"텔레그램 전송 실패: {resp.text}")


def main():
    start_dt, end_dt, date_str = get_target_date_range()
    logger.info(f"▶ 수집 시작: {date_str} (00:00 ~ 23:59)")

    items = fetch_all_pages(start_dt, end_dt)
    logger.info(f"전체 수집: {len(items)}건")

    if not items:
        y, m, d = date_str[:4], date_str[4:6], date_str[6:]
        send_telegram_message(
            f"📋 *나라장터 기술용역 사전규격공개*\n"
            f"📅 기준일: {y}-{m}-{d}\n"
            f"ℹ️ 해당일 등록 데이터가 없습니다."
        )
        return

    df = build_dataframe(items)
    logger.info(f"정렬 후 전체: {len(df)}건")

    # ★ 키워드 필터링
    df = filter_by_keywords(df)
    logger.info(f"최종 데이터: {len(df)}건")

    if df.empty:
        y, m, d = date_str[:4], date_str[4:6], date_str[6:]
        send_telegram_message(
            f"📋 *나라장터 기술용역 사전규격공개*\n"
            f"📅 기준일: {y}-{m}-{d}\n"
            f"ℹ️ 키워드 해당 데이터가 없습니다.\n"
            f"🔍 검색어: {', '.join(KEYWORDS)}"
        )
        return

    filepath = save_excel(df, date_str)
    send_telegram_file(filepath, date_str, len(df))
    logger.info("▶ 완료")


if __name__ == "__main__":
    main()
